#!/usr/bin/env python3
"""Generate the campaign-generation parity goldens (WP-14a).

Companion to `generate_goldens.py`, same contract: run it against the reference
toolkit in the sibling `amazon-agent` project, it imports the real
`campaign_model`, `build_campaigns` and `keyword_workbook` modules, replays the
create-mode scenarios its `selftest.py` pins plus the edge cases the TypeScript
port needs held down, and writes `{input, expected}` goldens into
`fixtures/golden/`.

    WIZARD_ADS_CAMPAIGN_REFERENCE_TOOLS=<path to amazon-agent/tools/amazon-campaign-builder> \\
        python3 fixtures/generate/generate_campaign_goldens.py

Regenerating is an operator step. CI replays the committed goldens with no
Python, no reference checkout and no network. The reference code is a SPEC,
never a dependency: nothing here is imported at runtime by the TypeScript
engine, and nothing is copied into it.

SYNTHETIC DATA ONLY. Every brand, product, SKU, ASIN, keyword and bid below is
invented for this file. A golden built from a real account, or carrying a real
strategy threshold, would be a client data leak with extra steps.

BMM IS OUT OF SCOPE (operator decision, 2026-08-14). The reference still builds
broad-match-modifier campaigns and this generator can still reach that code; no
scenario below asks for one, because the TypeScript port does not generate them
and a golden for a campaign type the port refuses would assert nothing. Reading
BMM is untouched: an account that already runs BMM campaigns is another work
package's problem, and this cut is about what we create.

Four goldens come out of this:

  campaigns.json          create-mode scenarios: config in, plan + bulk rows +
                          the Python-written workbook read back + preflight and
                          QA-gate verdicts out.
  campaign-preflight.json config-only cases, most of them deliberately invalid,
                          pinning every preflight issue and note string.
  campaign-validate.json  hand-built row sets, valid and deliberately broken,
                          pinning every QA-gate string.
  campaign-keywords.json  bucketed keyword sections in, campaign specs out.

`today` is pinned per run and recorded on every case, because the reference
reads the clock in three places (the default start date, the `Date` naming
variable, and the past-start-date gates). The TypeScript port takes it as an
argument instead, which is also what makes the port pure.
"""

from __future__ import annotations

import datetime as dt
import io
import json
import os
import sys
import tempfile
from contextlib import redirect_stdout
from pathlib import Path

HERE = Path(__file__).resolve().parent
GOLDEN_DIR = HERE.parent / "golden"

REFERENCE_ENV = "WIZARD_ADS_CAMPAIGN_REFERENCE_TOOLS"


def _load_reference() -> None:
    """Put the reference toolkit on the path, from the environment only.

    No absolute path to anybody's home directory is ever written into this
    repository, so the location arrives as an environment variable and a
    missing one is a clear instruction rather than an import traceback.
    """
    raw = os.environ.get(REFERENCE_ENV)
    if not raw:
        sys.exit(
            f"{REFERENCE_ENV} is not set.\n"
            "Point it at the reference toolkit directory (amazon-campaign-builder) in\n"
            "the sibling amazon-agent project, then re-run:\n"
            f"  {REFERENCE_ENV}=<path> python3 fixtures/generate/generate_campaign_goldens.py"
        )
    path = Path(raw).expanduser().resolve()
    if not (path / "campaign_model.py").is_file():
        sys.exit(f"{REFERENCE_ENV}={path} does not look like the reference toolkit (no campaign_model.py).")
    sys.path.insert(0, str(path))


_load_reference()
sys.path.insert(0, str(HERE))

import openpyxl  # noqa: E402

import build_campaigns as bc  # noqa: E402
import campaign_model as cm  # noqa: E402
import keyword_workbook as kwb  # noqa: E402
import serialize_campaigns as S  # noqa: E402

SCHEMA = "wizard-ads.parity.v1"
SOURCE = "amazon-agent/tools/amazon-campaign-builder (read-only reference)"

#: The clock, pinned once per run and recorded on every case.
TODAY = dt.date.today()

#: A date that is always in the future, for the cases that need a start date the
#: reference's past-date gates accept no matter when the goldens were made.
FUTURE = "2099-01-15"
#: And one that is always in the past, to pin the gate that rejects it.
PAST = "2020-01-01"

TMP = Path(tempfile.mkdtemp(prefix="wizard-ads-campaign-goldens-"))


# ---------------------------------------------------------------------------
# scenario configs. Invented brands, invented products, invented ASINs.


def base(**overrides) -> dict:
    cfg = {
        "client": "Sample Tools Co",
        "brand": "Sample",
        "marketplace": "US",
        "defaults": {"daily_budget": 10.0, "keyword_bid": 0.5, "state": "paused"},
        "campaigns": [],
    }
    cfg.update(overrides)
    return cfg


#: A. The reference selftest's LEGACY-preset scenario, verbatim.
LEGACY_PRESET = base(
    client="Sample Legacy",
    naming={"preset": "LEGACY", "suffix": "EW"},
    campaigns=[
        {"campaign_type": "SKW", "product_name": "Widget", "target_descriptor": "generic",
         "sku": ["SKU-1"], "keywords": ["red widget", "blue widget"]},
        {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "long-tail",
         "sku": ["SKU-1"], "keywords": ["red widget for kitchen", "red widget for office"]},
        {"campaign_type": "Phrase", "product_name": "Widget", "target_descriptor": "widget",
         "sku": ["SKU-1"], "keywords": ["widget"]},
        {"campaign_type": "Auto", "product_name": "Widget", "target_descriptor": "auto",
         "sku": ["SKU-1"]},
        {"campaign_type": "PAT", "product_name": "Widget", "target_descriptor": "competitors",
         "sku": ["SKU-1"], "target_asins": ["B000000001", "B000000002"]},
    ],
)

#: B. The reference selftest's EW-preset scenario (the default preset), verbatim.
EW_PRESET = base(
    client="Sample EW",
    campaigns=[
        {"campaign_type": "SKW", "product_name": "Widget", "sku": ["SKU-1"],
         "keywords": ["red widget"], "top_of_search_placement": 55,
         "child_state": "enabled"},
        {"campaign_type": "SKW", "campaign_purpose": "SHIELD", "product_name": "Widget",
         "sku": ["SKU-1"], "keywords": ["sample widget"]},
        {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "long-tail",
         "sku": ["SKU-1"], "keywords": ["red widget for kitchen", "red widget for office"]},
        {"campaign_type": "Auto", "product_name": "Widget", "sku": ["SKU-1"]},
        {"campaign_type": "PAT", "campaign_purpose": "SELF_TARGETING", "match_type": "ASIN_EXPANDED",
         "product_name": "Widget", "sku": ["SKU-1"], "target_asins": ["B000000009"]},
        {"campaign_type": "PAT", "product_name": "Widget", "sku": ["SKU-1"],
         "target_asins": ["B000000001"], "bidding_strategy": "Up and down"},
        {"campaign_type": "PAT", "goal": "Brand", "product_name": "Widget",
         "target_descriptor": "category", "sku": ["SKU-1"],
         "target_categories": ["123456"],
         "negative_target_asins": ["B000000002", "B000000003"]},
    ],
)

#: Keyword chunking: one campaign per N keywords, disambiguated by the counter.
TRANSPOSE_CHUNKS = base(
    client="Sample Transpose",
    naming={"preset": "LEGACY", "variable_order": ["Goal", "SP", "MatchType", "ProductName",
                                                   "TargetDescriptor", "Counter", "EW"]},
    campaigns=[
        {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "long-tail",
         "sku": ["SKU-1"], "transpose_keywords": True, "keywords_per_campaign": 2,
         "keywords": ["widget for kitchen", "widget for office", "widget for garage",
                      "widget for shed", "widget for attic"]},
    ],
)

#: Negatives at both levels and in both match types.
NEGATIVES_BOTH_LEVELS = base(
    client="Sample Negatives",
    campaigns=[
        {"campaign_type": "Phrase", "product_name": "Widget", "target_descriptor": "widget",
         "sku": ["SKU-1"], "keywords": ["red widget", "blue widget"],
         "negative_keywords": ["free widget", "widget repair manual"],
         "negative_match_type": "NEGATIVE_PHRASE"},
        {"campaign_type": "Phrase", "product_name": "Widget", "target_descriptor": "holder",
         "sku": ["SKU-1"], "keywords": ["widget holder"],
         "negative_keywords": ["used widget"], "negative_level": "campaign"},
    ],
)

#: ProductName and TargetDescriptor swap places in the name.
SWAP_NAME_ORDER = base(
    client="Sample Swap",
    naming={"preset": "LEGACY"},
    campaigns=[
        {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "kitchen",
         "sku": ["SKU-1"], "keywords": ["widget for kitchen"], "swap_name_order": True},
        {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "office",
         "sku": ["SKU-1"], "keywords": ["widget for office"]},
    ],
)

#: Vendor mode blanks the SKU column; several products advertise in one ad group.
VENDOR_MULTI_PRODUCT = base(
    client="Sample Vendor",
    vendor_central_mode=True,
    campaigns=[
        {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "bundle",
         "sku": "SKU-1, SKU-2", "asin": "B000000011, B000000012",
         "keywords": ["widget bundle", "widget kit"]},
    ],
)

#: Auto targeting groups with their own bids and states, plus file-wide placements.
AUTO_GROUP_OVERRIDES = base(
    client="Sample Auto",
    defaults={"daily_budget": 12.0, "keyword_bid": 0.45, "state": "paused",
              "start_date": FUTURE, "portfolio_id": "5551234", "site_restriction": "Amazon Business",
              "top_of_search_placement": 30, "rest_of_search_placement": 10,
              "product_pages_placement": 0},
    campaigns=[
        {"campaign_type": "Auto", "product_name": "Widget", "sku": ["SKU-1"],
         "auto_close_match_bid": 0.6, "auto_close_match_state": "enabled",
         "auto_loose_match_bid": 0.3, "auto_loose_match_state": "paused",
         "auto_substitutes_bid": 0.35,
         "auto_complements_state": "paused"},
    ],
)

#: SKW fan-out where the keyword is NOT in the name, so the counter carries it.
SKW_COUNTER_NAMES = base(
    client="Sample Counter",
    naming={"preset": "LEGACY", "variable_order": ["Goal", "SP", "MatchType", "ProductName",
                                                   "TargetDescriptor", "Counter", "EW"]},
    campaigns=[
        {"campaign_type": "SKW", "product_name": "Widget", "target_descriptor": "rank",
         "sku": ["SKU-1"], "skw_include_keyword_in_name": False,
         "keywords": ["red widget", "blue widget", "green widget"]},
    ],
)

#: Every naming variable the reference knows, including the date and the two
#: custom slots, on a non-default delimiter.
CUSTOM_NAMING_TOKENS = base(
    client="Sample Naming",
    naming={
        "variable_order": ["Goal", "AdType", "CampaignType", "MatchType", "TriggerWord",
                           "ProductName", "TargetDescriptor", "Keyword", "Counter",
                           "CampCounter", "Date", "Custom1", "Custom2", "EW"],
        "delimiter": "_",
        "suffix": "SFX",
        "custom1_value": "C1",
        "custom2_value": "C2",
    },
    campaigns=[
        {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "kitchen",
         "sku": ["SKU-1"], "keywords": ["widget for kitchen"]},
        {"campaign_type": "SKW", "product_name": "Widget", "sku": ["SKU-1"],
         "keywords": ["red widget"]},
    ],
)

#: Per-campaign placement overrides, including an explicit zero that must beat a
#: non-zero file default, and a child state that differs from the campaign state.
PLACEMENT_OVERRIDES = base(
    client="Sample Placements",
    defaults={"daily_budget": 10.0, "keyword_bid": 0.5, "state": "paused",
              "top_of_search_placement": 40, "rest_of_search_placement": 15,
              "product_pages_placement": 5},
    campaigns=[
        {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "override",
         "sku": ["SKU-1"], "keywords": ["widget for kitchen"],
         "top_of_search_placement": 0, "product_pages_placement": 900,
         "child_state": "enabled"},
        {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "inherit",
         "sku": ["SKU-1"], "keywords": ["widget for office"]},
    ],
)


def drop_bmm_modifier(specs: list) -> list:
    """Strip the one spec key that only ever meant something for BMM.

    `sections_to_campaign_specs` stamps `bmm_modifier` on every discovery spec,
    `True` for a BMM root and `False` for a Phrase one. With BMM out of scope
    the key is always `False`, so the port does not carry it and the golden
    should not assert it. Normalizing here keeps that decision visible instead
    of leaving the port quietly ignoring an input field.
    """
    return [{k: v for k, v in spec.items() if k != "bmm_modifier"} for spec in specs]


def keyword_file_specs() -> list:
    """C. The reference selftest's keyword-file scenario.

    The specs come out of the real workbook parser rather than being written by
    hand, so the golden's input is a shape the reference actually produces.
    """
    path = TMP / "campaign_structure_fixture.xlsx"
    _write_campaign_structure_fixture(path)
    return drop_bmm_modifier(
        kwb.parse_keyword_workbook(str(path), product_name="Widget", sku=["SKU-1"])
    )


def _write_campaign_structure_fixture(path: Path) -> None:
    """A synthetic '5. Campaign Structure' tab, the shape the reference parses.

    Structurally identical to the reference selftest's fixture; the keywords and
    ASINs are invented here as they are there.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "5. Campaign Structure"

    def section(row, col, title, kind, label=""):
        ws.cell(row, col, title)
        hdr_row = row + (2 if label else 1)
        if label:
            ws.cell(row + 1, col, label)
        if kind == "keywords":
            ws.cell(hdr_row, col, "Keyword")
            ws.cell(hdr_row, col + 1, "Search Volume")
        else:
            ws.cell(hdr_row, col, "ASINs")
            ws.cell(hdr_row, col + 1, "Brand")
        return hdr_row

    h1 = section(1, 1, "Rank-SKW", "keywords")
    for i, (kw, sv) in enumerate([("red widget", 2000), ("blue widget", 1500)]):
        ws.cell(h1 + 1 + i, 1, kw)
        ws.cell(h1 + 1 + i, 2, sv)
    ws.cell(h1 + 3, 1, "Sum")

    h2 = section(1, 4, "Shield-SKW", "keywords")
    ws.cell(h2 + 1, 4, "sample widget")
    ws.cell(h2 + 1, 5, 500)
    ws.cell(h2 + 3, 4, "Sum")

    h3 = section(10, 1, "Long-Tails (Halo)", "keywords")
    for i, kw in enumerate(["red widget for kitchen", "red widget for office"]):
        ws.cell(h3 + 1 + i, 1, kw)
        ws.cell(h3 + 1 + i, 2, 300)
    ws.cell(h3 + 4, 1, "Sum")

    # One discovery column, not two. The reference's scanner reads the label
    # above the header to tell a BMM root from a Phrase root; with BMM out of
    # scope there is only ever a Phrase root to find.
    h4 = section(20, 1, "Discovery-Root Keywords", "keywords", label="Phrase Root")
    ws.cell(h4 + 1, 1, "widget")
    ws.cell(h4 + 1, 2, 5000)
    ws.cell(h4 + 2, 1, "widget accessory")
    ws.cell(h4 + 2, 2, 4000)
    ws.cell(h4 + 4, 1, "Sum")

    h6 = section(30, 1, "PAT (Stronger)", "asins")
    ws.cell(h6 + 1, 1, "B000000001")
    ws.cell(h6 + 1, 2, "Rival One")
    ws.cell(h6 + 3, 1, "Sum")

    h7 = section(30, 4, "PAT (Weaker)", "asins")
    ws.cell(h7 + 1, 4, "B000000002")
    ws.cell(h7 + 1, 5, "Rival Two")
    ws.cell(h7 + 3, 4, "Sum")

    wb.save(path)


# ---------------------------------------------------------------------------
# running the reference


def _captured(fn, *args) -> tuple[int, str]:
    buf = io.StringIO()
    with redirect_stdout(buf):
        rc = fn(*args)
    return rc, buf.getvalue()


def _lines_tagged(text: str, tag: str) -> list:
    """Pull one class of reference report line out of captured stdout."""
    prefix = f"  [{tag}]"
    out = []
    for line in text.splitlines():
        if line.startswith(prefix):
            out.append(line[len(prefix):].strip())
    return out


def preflight_verdict(loaded: dict) -> dict:
    rc, text = _captured(bc.preflight, loaded)
    return {
        "ready": rc == 0,
        "issues": _lines_tagged(text, "MISSING"),
        "notes": _lines_tagged(text, "NOTE"),
    }


def validate_verdict(loaded: dict, xlsx: Path) -> dict:
    rc, text = _captured(bc.validate, loaded, str(xlsx))
    return {
        "pass": rc == 0,
        "fails": _lines_tagged(text, "FAIL"),
        "warns": _lines_tagged(text, "WARN"),
    }


def write_workbook(rows: list, xlsx: Path) -> None:
    """The reference's own workbook writer, with the clock taken out.

    `build_campaigns.build` is these seven lines plus a `date.today()` default
    that would make the golden depend on the day it was generated; the rows are
    built above with `today` injected instead, so this writes exactly what the
    reference would have written on that day.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cm.SHEET_NAMES["SP"]
    ws.append(cm.COLUMNS["SP"])
    for row in rows:
        ws.append([row[c] for c in cm.COLUMNS["SP"]])
    for i, col in enumerate(cm.COLUMNS["SP"], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = max(len(col) + 2, 18)
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx)


def read_workbook(xlsx: Path) -> dict:
    """Read the written workbook back, exactly as the QA gate does."""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[cm.SHEET_NAMES["SP"]]
    grid = [[S.cell(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
    return {
        "sheetNames": list(wb.sheetnames),
        "header": grid[0],
        "rows": grid[1:],
    }


def generate_all_with_today(loaded: dict) -> list:
    """`build_campaigns.generate_all`, with the clock passed in.

    The reference calls `generate_campaigns(form, naming)` and lets the `Date`
    naming variable read `date.today()`. Passing `TODAY` explicitly is the only
    difference, and it is what lets the port stay a pure function.
    """
    campaigns = []
    for form in bc.campaign_forms(loaded):
        campaigns.extend(cm.generate_campaigns(form, loaded["naming"], TODAY))
    return campaigns


def build_defaults(loaded: dict) -> dict:
    """The defaults block `build_campaigns.build` hands to the row builder."""
    return {**loaded["defaults"],
            "vendor_central_mode": bool(loaded["defaults"].get("vendor_central_mode")
                                        or loaded.get("vendor_central_mode"))}


def load(cfg: dict) -> dict:
    """Run a raw config through the reference's own naming-preset resolution."""
    path = TMP / "config.json"
    path.write_text(json.dumps(cfg))
    return bc.load_config(str(path))


# ---------------------------------------------------------------------------
# case builders


def campaign_cases() -> list:
    scenarios = [
        ("legacy_preset", LEGACY_PRESET),
        ("ew_preset", EW_PRESET),
        ("keyword_file", base(client="Sample Keyword File", campaigns=keyword_file_specs())),
        ("transpose_chunks", TRANSPOSE_CHUNKS),
        ("negatives_both_levels", NEGATIVES_BOTH_LEVELS),
        ("swap_name_order", SWAP_NAME_ORDER),
        ("vendor_multi_product", VENDOR_MULTI_PRODUCT),
        ("auto_group_overrides", AUTO_GROUP_OVERRIDES),
        ("skw_counter_names", SKW_COUNTER_NAMES),
        ("custom_naming_tokens", CUSTOM_NAMING_TOKENS),
        ("placement_overrides", PLACEMENT_OVERRIDES),
    ]

    cases = []
    for name, cfg in scenarios:
        loaded = load(cfg)
        campaigns = generate_all_with_today(loaded)
        rows = cm.build_bulk_rows(campaigns, build_defaults(loaded), today=TODAY)
        xlsx = TMP / f"{name}.xlsx"
        write_workbook(rows, xlsx)
        cases.append({
            "name": f"generate:{name}",
            "input": {"config": S.config(loaded), "today": TODAY.isoformat()},
            "expected": {
                "campaigns": [S.campaign(c) for c in campaigns],
                "rows": [S.bulk_row(r, cm.SP_COLUMNS) for r in rows],
                "workbook": read_workbook(xlsx),
                "preflight": preflight_verdict(loaded),
                "validate": validate_verdict(loaded, xlsx),
            },
        })
    return cases


def preflight_cases() -> list:
    """Config-only cases. Most are deliberately invalid: the reference's issue
    and note strings are the contract, so every one of them is pinned."""
    # No unknown-campaign-type case here on purpose: the reference's message for
    # one enumerates its own type list, which still contains BMM. The port's
    # list does not, so that single string cannot be held to parity and is
    # covered by a unit test in the package instead.
    scenarios = [
        ("missing_product_name", base(campaigns=[
            {"campaign_type": "Halo", "sku": ["SKU-1"], "keywords": ["widget"]}])),
        ("missing_sku", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "keywords": ["widget"]}])),
        ("vendor_without_asin", base(vendor_central_mode=True, campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"]}])),
        ("missing_keywords", base(campaigns=[
            {"campaign_type": "Phrase", "product_name": "Widget", "sku": ["SKU-1"]}])),
        ("pat_without_targets", base(campaigns=[
            {"campaign_type": "PAT", "product_name": "Widget", "sku": ["SKU-1"]}])),
        ("pat_bad_asins", base(campaigns=[
            {"campaign_type": "PAT", "product_name": "Widget", "sku": ["SKU-1"],
             "target_asins": ["nope", "B000000001"]}])),
        ("pat_bad_categories", base(campaigns=[
            {"campaign_type": "PAT", "product_name": "Widget", "sku": ["SKU-1"],
             "target_categories": ["kitchen", "123456"]}])),
        ("bad_negative_asins", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "negative_target_asins": ["short"]}])),
        ("bad_goal", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "goal": "Growth"}])),
        ("unusual_goal", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "goal": "Discovery"}])),
        ("bad_match_type", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "match_type": "NEAR"}])),
        ("bad_purpose", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "campaign_purpose": "PROFIT"}])),
        ("bad_bidding_strategy", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "bidding_strategy": "Aggressive"}])),
        ("bidding_override_note", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "bidding_strategy": "Up and down"}])),
        ("bad_states", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "state": "archived", "child_state": "running"}])),
        ("bid_out_of_range", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "keyword_bid": 0.01}])),
        ("budget_below_minimum", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "daily_budget": 0.5}])),
        ("placement_out_of_range", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "top_of_search_placement": 901}])),
        ("bad_negative_settings", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "negative_match_type": "NEGATIVE_BROAD",
             "negative_level": "profile"}])),
        ("transpose_without_size", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["a widget", "b widget"], "transpose_keywords": True}])),
        ("fan_out_collision", base(
            naming={"preset": "LEGACY"},
            campaigns=[
                {"campaign_type": "Halo", "product_name": "Widget", "target_descriptor": "long",
                 "sku": ["SKU-1"], "transpose_keywords": True, "keywords_per_campaign": 1,
                 "keywords": ["a widget", "b widget"]}])),
        ("discovery_without_negatives", base(campaigns=[
            {"campaign_type": "Phrase", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"]}])),
        ("self_targeting_without_negatives", base(campaigns=[
            {"campaign_type": "PAT", "campaign_purpose": "SELF_TARGETING",
             "match_type": "ASIN_EXPANDED", "product_name": "Widget", "sku": ["SKU-1"],
             "target_asins": ["B000000009"]}])),
        ("past_start_date", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "start_date": PAST}])),
        ("malformed_start_date", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "start_date": "15/01/2099"}])),
        ("enabled_state_note", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "state": "enabled", "start_date": FUTURE}])),
        ("ad_group_name_equals_campaign_name", base(
            naming={"variable_order": ["MatchType", "ProductName"], "delimiter": " | "},
            campaigns=[
                {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
                 "keywords": ["widget"]}])),
        ("missing_client_and_campaigns", {"marketplace": "", "defaults": {}, "campaigns": []}),
        ("ready", base(campaigns=[
            {"campaign_type": "Halo", "product_name": "Widget", "sku": ["SKU-1"],
             "keywords": ["widget"], "start_date": FUTURE}])),
    ]

    cases = []
    for name, cfg in scenarios:
        loaded = load(cfg)
        cases.append({
            "name": f"preflight:{name}",
            "input": {"config": S.config(loaded), "today": TODAY.isoformat()},
            "expected": preflight_verdict(loaded),
        })
    return cases


def _row(**kw) -> dict:
    row = {c: "" for c in cm.SP_COLUMNS}
    row["Product"] = "Sponsored Products"
    row["Operation"] = "Create"
    row.update(kw)
    return row


def validate_cases() -> list:
    """Hand-built row sets straight into the QA gates.

    Building the rows by hand rather than generating them is the point: the
    gates exist to catch a file that the generator would never produce, so a
    golden made only from generated rows would never exercise them.
    """
    today_compact = TODAY.isoformat().replace("-", "")
    healthy = [
        _row(Entity="Campaign", **{"Campaign ID": "tmp-1", "Campaign Name": "C1",
             "Start Date": today_compact, "Targeting Type": "MANUAL", "State": "paused",
             "Daily Budget": 10.0, "Bidding Strategy": "Dynamic bids - down only"}),
        _row(Entity="Ad Group", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused",
             "Ad Group Default Bid": 0.5}),
        _row(Entity="Product Ad", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused", "SKU": "SKU-1"}),
        _row(Entity="Keyword", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused", "Bid": 0.5,
             "Keyword Text": "widget", "Match Type": "exact"}),
    ]

    numeric_ids = [
        _row(Entity="Campaign", **{"Campaign ID": "1", "Campaign Name": "C1",
             "Start Date": today_compact, "Targeting Type": "MANUAL", "State": "paused",
             "Daily Budget": 10.0, "Bidding Strategy": "Dynamic bids - down only"}),
        _row(Entity="Ad Group", **{"Campaign ID": "1", "Ad Group ID": "2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused",
             "Ad Group Default Bid": 0.5}),
        _row(Entity="Product Ad", **{"Campaign ID": "1", "Ad Group ID": "2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused", "SKU": "SKU-1"}),
        _row(Entity="Keyword", **{"Campaign ID": "1", "Ad Group ID": "2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused", "Bid": 0.5,
             "Keyword Text": "widget", "Match Type": "exact"}),
    ]

    duplicate_names = healthy + [
        _row(Entity="Campaign", **{"Campaign ID": "tmp-3", "Campaign Name": "C1",
             "Start Date": today_compact, "Targeting Type": "MANUAL", "State": "paused",
             "Daily Budget": 10.0, "Bidding Strategy": "Dynamic bids - down only"}),
        _row(Entity="Ad Group", **{"Campaign ID": "tmp-3", "Ad Group ID": "tmp-4",
             "Campaign Name": "C1", "Ad Group Name": "AG2", "State": "paused",
             "Ad Group Default Bid": 0.5}),
        _row(Entity="Product Ad", **{"Campaign ID": "tmp-3", "Ad Group ID": "tmp-4",
             "Campaign Name": "C1", "Ad Group Name": "AG2", "State": "paused", "SKU": "SKU-2"}),
        _row(Entity="Keyword", **{"Campaign ID": "tmp-3", "Ad Group ID": "tmp-4",
             "Campaign Name": "C1", "Ad Group Name": "AG2", "State": "paused", "Bid": 0.5,
             "Keyword Text": "widget", "Match Type": "exact"}),
    ]

    duplicate_keyword = healthy + [
        _row(Entity="Keyword", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused", "Bid": 0.5,
             "Keyword Text": "Widget", "Match Type": "exact"}),
    ]

    orphans = [
        _row(Entity="Keyword", **{"Campaign ID": "tmp-9", "Ad Group ID": "tmp-8",
             "Campaign Name": "C9", "Ad Group Name": "AG9", "State": "paused", "Bid": 0.5,
             "Keyword Text": "widget", "Match Type": "exact"}),
    ]

    bad_values = [
        _row(Entity="Campaign", Operation="Update", **{"Campaign ID": "tmp-1",
             "Campaign Name": "C" * 200, "Start Date": "2026-01-01", "Targeting Type": "SMART",
             "State": "archived", "Daily Budget": 0.5, "Bidding Strategy": "Down only"}),
        _row(Entity="Bidding Adjustment", **{"Campaign ID": "tmp-1", "Campaign Name": "C1",
             "Placement": "Placement Rest of Search", "Percentage": 950}),
        _row(Entity="Ad Group", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "A" * 300, "State": "paused",
             "Ad Group Default Bid": 5000.0}),
        _row(Entity="Product Ad", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused"}),
        _row(Entity="Keyword", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused", "Bid": 0.0,
             "Keyword Text": "widget", "Match Type": "EXACT"}),
        _row(Entity="Negative Keyword", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused",
             "Keyword Text": "one two three four five six seven eight nine ten eleven",
             "Match Type": "NEGATIVE_EXACT"}),
        _row(Entity="Product Targeting", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused", "Bid": 0.5,
             "Product Targeting Expression": 'asin="nope"'}),
    ]

    no_targets = [
        _row(Entity="Campaign", **{"Campaign ID": "tmp-1", "Campaign Name": "C1",
             "Start Date": today_compact, "Targeting Type": "MANUAL", "State": "paused",
             "Daily Budget": 10.0, "Bidding Strategy": "Dynamic bids - down only"}),
        _row(Entity="Ad Group", **{"Campaign ID": "tmp-1", "Ad Group ID": "tmp-2",
             "Campaign Name": "C1", "Ad Group Name": "AG1", "State": "paused",
             "Ad Group Default Bid": 0.5}),
    ]

    scenarios = [
        ("healthy", healthy),
        ("numeric_temp_ids", numeric_ids),
        ("duplicate_campaign_names", duplicate_names),
        ("duplicate_keyword_in_ad_group", duplicate_keyword),
        ("orphan_rows", orphans),
        ("bad_values", bad_values),
        ("campaign_without_targets", no_targets),
    ]

    cfg = load(base(campaigns=[]))
    cases = []
    for name, rows in scenarios:
        xlsx = TMP / f"validate_{name}.xlsx"
        write_workbook(rows, xlsx)
        cases.append({
            "name": f"validate:{name}",
            "input": {
                "rows": [S.bulk_row(r, cm.SP_COLUMNS) for r in rows],
                "today": TODAY.isoformat(),
            },
            "expected": validate_verdict(cfg, xlsx),
        })
    return cases


def keyword_cases() -> list:
    """Bucketed keyword sections to campaign specs.

    The section shapes are the ones the reference's own workbook scanner emits;
    the first case is that scanner's live output on the synthetic workbook, so
    the hand-written cases that follow cannot drift from a shape the reference
    never produces.
    """
    path = TMP / "campaign_structure_fixture.xlsx"
    if not path.exists():
        _write_campaign_structure_fixture(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    scanned = kwb.scan_campaign_structure_sections(wb["5. Campaign Structure"])

    hand_built = [
        {"campaign_type": "SKW", "campaign_purpose": "RANK_SKW", "label": "", "kind": "keywords",
         "values": [("red widget", 2000), ("blue widget", 1500)]},
        {"campaign_type": "SKW", "campaign_purpose": "SHIELD", "label": "", "kind": "keywords",
         "values": [("sample widget", 500)]},
        {"campaign_type": "Halo", "campaign_purpose": "HALO", "label": "long-tail",
         "kind": "keywords", "values": [("widget for kitchen", 300), ("widget for office", 250)]},
        {"campaign_type": "Phrase", "campaign_purpose": "DISCOVERY", "label": "Phrase Root",
         "kind": "keywords", "values": [("widget", 5000), ("widget holder", 900)]},
        {"campaign_type": "Phrase", "campaign_purpose": "SHIELD", "label": "Phrase Brand",
         "kind": "keywords", "values": [("sample widget holder", 400)]},
        {"campaign_type": "PAT", "campaign_purpose": "DISCOVERY", "label": "Stronger",
         "kind": "asins", "values": [("b000000001", "Rival One"), ("B000000002", "Rival Two")]},
    ]

    scenarios = [
        ("scanned_workbook", scanned, {"product_name": "Widget", "sku": ["SKU-1"], "asin": None}),
        ("hand_built_buckets", hand_built,
         {"product_name": "Widget", "sku": ["SKU-1", "SKU-2"], "asin": ["B000000011"]}),
        ("no_product_identity", hand_built[:1], {"product_name": "", "sku": None, "asin": None}),
        ("empty", [], {"product_name": "Widget", "sku": ["SKU-1"], "asin": None}),
    ]

    cases = []
    for name, sections, kwargs in scenarios:
        specs = drop_bmm_modifier(kwb.sections_to_campaign_specs(sections, **kwargs))
        cases.append({
            "name": f"sections_to_specs:{name}",
            "input": {
                "sections": [{
                    "campaignType": s["campaign_type"],
                    "campaignPurpose": s["campaign_purpose"],
                    "label": s["label"],
                    "kind": s["kind"],
                    "values": [[v[0], v[1]] for v in s["values"]],
                } for s in sections],
                "productName": kwargs["product_name"],
                "sku": list(kwargs["sku"] or []),
                "asin": list(kwargs["asin"] or []),
            },
            "expected": [S.spec(s) for s in specs],
        })
    return cases


# ---------------------------------------------------------------------------


def write(module: str, cases: list) -> None:
    GOLDEN_DIR.mkdir(parents=True, exist_ok=True)
    path = GOLDEN_DIR / f"{module}.json"
    payload = {
        "schema": SCHEMA,
        "module": module,
        "source": SOURCE,
        "synthetic": True,
        "today": TODAY.isoformat(),
        "columns": list(cm.SP_COLUMNS),
        "sheetName": cm.SHEET_NAMES["SP"],
        "cases": cases,
    }
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False, allow_nan=False, sort_keys=False)
        fh.write("\n")
    print(f"{path.relative_to(GOLDEN_DIR.parent.parent)}: {len(cases)} cases")


def main() -> int:
    write("campaigns", campaign_cases())
    write("campaign-preflight", preflight_cases())
    write("campaign-validate", validate_cases())
    write("campaign-keywords", keyword_cases())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
